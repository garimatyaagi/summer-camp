from openpyxl import load_workbook

wb = load_workbook("/Users/garimatyagi/Coding/summer-camp/Summer_Club_Budget.xlsx")
ws = wb["Simulator"]

# Update inputs with realistic Bangalore pricing
updates = {
    # (row, col): value
    (6, 3): 18,        # Kids (unchanged)
    (7, 3): 21000,     # Early bird (unchanged)
    (8, 3): 24500,     # Regular (unchanged)
    (9, 3): 0.6,       # Early bird % (unchanged)
    (11, 3): 45000,    # Venue rent — half-day usage, gated space, South Bangalore
    (12, 3): 90000,    # Venue deposit — 2 months (refundable)
    (13, 3): 95000,    # Staff total/month: instructor 1 (₹35K) + instructor 2 (₹30K) + helper (₹18K) + first aid part-time (₹12K)
    (14, 3): 3500,     # Materials per child (7 weeks) — quality supplies, not the cheapest
    (15, 3): 55,       # Snacks — good fruit + juice, not canteen grade
    (16, 3): 1000,     # Per-child ops — quality t-shirt (₹450) + scrapbook (₹300) + cert (₹150) + tag (₹100)
    (17, 3): 15000,    # Marketing — flyers, posters, Instagram ads, WhatsApp campaign
    (18, 3): 15000,    # Insurance + first aid supplies + safety equipment
    (19, 3): 0.1,      # Buffer (unchanged)
}

# Update notes
notes = {
    11: "Half-day (9-1pm), gated campus, AC, South BLR",
    12: "2 months deposit (refundable after camp)",
    13: "Instructor ₹35K + ₹30K + Helper ₹18K + First aid ₹12K",
    14: "₹500/child/week avg across all 7 themes",
    15: "Quality fruit + juice, ₹55/child/day",
    16: "T-shirt ₹450 + Scrapbook ₹300 + Cert ₹150 + Tag ₹100",
    17: "Flyers, posters, Instagram boost, WhatsApp tools",
    18: "Group activity insurance + first aid kit + safety gear",
}

for (r, c), v in updates.items():
    ws.cell(r, c, v)

for r, note in notes.items():
    from openpyxl.styles import Font
    ws.cell(r, 4, note).font = Font(color="999999", size=9, name="Arial", italic=True)

# Also update the Costs sheet with realistic numbers
ws2 = wb["Costs"]
cost_updates = {
    (6, 2): 45000,   # Venue rent
    (7, 2): 90000,   # Deposit
    (10, 2): 35000,  # Instructor 1
    (11, 2): 30000,  # Instructor 2
    (12, 2): 18000,  # Helper
    (13, 2): 12000,  # First aid
    (16, 2): 550,    # Art supplies per child
    (17, 2): 500,    # Science per child
    (18, 2): 450,    # Cooking per child
    (19, 2): 300,    # Nature per child
    (20, 2): 600,    # Robot per child
    (21, 2): 450,    # Market per child
    (22, 2): 700,    # Exhibition per child
    (25, 2): 55,     # Snacks per child per day
    (28, 2): 5000,   # Flyers
    (29, 2): 10000,  # Instagram ads
    (32, 2): 12000,  # Insurance
    (33, 2): 5000,   # First aid kit
    (34, 2): 200,    # Certificates
    (35, 2): 120,    # Name tags
    (36, 2): 450,    # T-shirts
    (37, 2): 300,    # Scrapbooks
}
for (r, c), v in cost_updates.items():
    ws2.cell(r, c, v)

# Update Revenue sheet assumptions too
ws1 = wb["Revenue"]
ws1.cell(6, 2, 21000)
ws1.cell(7, 2, 24500)

wb.save("/Users/garimatyagi/Coding/summer-camp/Summer_Club_Budget.xlsx")
print("Updated with realistic Bangalore costs")
