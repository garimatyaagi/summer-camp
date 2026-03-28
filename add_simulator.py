from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList

wb = load_workbook("/Users/garimatyagi/Coding/summer-camp/Summer_Club_Budget.xlsx")

blue_font = Font(color="0000FF", size=11, name="Arial")
black_font = Font(color="000000", size=11, name="Arial")
bold_black = Font(color="000000", size=11, name="Arial", bold=True)
bold_white = Font(color="FFFFFF", size=11, name="Arial", bold=True)
header_font = Font(color="FFFFFF", size=12, name="Arial", bold=True)
section_font = Font(color="2B5797", size=13, name="Arial", bold=True)
title_font = Font(color="2B5797", size=16, name="Arial", bold=True)
subtitle_font = Font(color="6B7D8F", size=11, name="Arial")
big_num = Font(color="2B5797", size=28, name="Arial", bold=True)
big_green = Font(color="006100", size=28, name="Arial", bold=True)
big_red = Font(color="CC0000", size=28, name="Arial", bold=True)
label_font = Font(color="6B7D8F", size=10, name="Arial")

header_fill = PatternFill("solid", fgColor="2B5797")
yellow_fill = PatternFill("solid", fgColor="FFFF00")
green_fill = PatternFill("solid", fgColor="E2EFDA")
red_fill = PatternFill("solid", fgColor="FCE4EC")
light_fill = PatternFill("solid", fgColor="F0F4F9")
input_fill = PatternFill("solid", fgColor="FFF9E6")

inr = '₹#,##0;(₹#,##0);"-"'
pct = '0.0%'
num = '#,##0'
thick = Side(style="medium", color="2B5797")
double_border = Border(top=thick, bottom=Side(style="double", color="2B5797"))

ws = wb.create_sheet("Simulator", 0)  # Make it first sheet

for i, w in enumerate([3, 35, 22, 22, 5, 22, 22, 5, 35, 22], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── TITLE ──
ws.merge_cells("B1:G1")
ws.cell(1, 2, "THE SUMMER CLUB — SCENARIO SIMULATOR").font = title_font
ws.cell(2, 2, "Change the yellow cells to see how your numbers shift instantly").font = subtitle_font

# ── INPUT SECTION ──
r = 4
ws.cell(r, 2, "YOUR INPUTS").font = section_font
for c in range(2, 8):
    ws.cell(r, c).border = Border(bottom=thick)

inputs = [
    (6, "Number of kids", 18, num, "Change this to see break-even"),
    (7, "Early bird price", 21000, inr, "Price for first 25 families"),
    (8, "Regular price", 24500, inr, "Full price after early bird"),
    (9, "% paying early bird", 0.6, pct, "What portion gets the discount"),
    (10, "", None, None, None),
    (11, "Venue rent (per month)", 20000, inr, "2 months needed"),
    (12, "Venue deposit (refundable)", 20000, inr, "Usually 1 month"),
    (13, "Staff cost (per month, total)", 55000, inr, "2 instructors + helper + first aid"),
    (14, "Materials per child (7 weeks)", 2550, inr, "All 7 weeks combined"),
    (15, "Snacks per child per day", 30, inr, "Fruit + juice"),
    (16, "Per-child ops (tshirt, cert, etc)", 800, inr, "T-shirt + scrapbook + cert + tag"),
    (17, "Marketing budget", 8000, inr, "Flyers + Instagram"),
    (18, "Insurance + first aid kit", 11000, inr, "Group activity insurance"),
    (19, "Buffer %", 0.1, pct, "Safety margin"),
]

for row, label, val, fmt, note in inputs:
    ws.cell(row, 2, label).font = black_font
    if val is not None:
        cell = ws.cell(row, 3, val)
        cell.font = blue_font
        cell.number_format = fmt
        cell.fill = yellow_fill
        cell.alignment = Alignment(horizontal="center")
    if note:
        ws.cell(row, 4, note).font = Font(color="999999", size=9, name="Arial", italic=True)

# ── CALCULATED RESULTS ──
r = 21
ws.cell(r, 2, "RESULTS").font = section_font
for c in range(2, 8):
    ws.cell(r, c).border = Border(bottom=thick)

# Revenue calc
results = [
    (23, "REVENUE", None, None, True),
    (24, "Early bird kids", "=ROUND(C6*C9,0)", num),
    (25, "Regular kids", "=C6-C24", num),
    (26, "Early bird revenue", "=C24*C7", inr),
    (27, "Regular revenue", "=C25*C8", inr),
    (28, "Total Revenue", "=C26+C27", inr, True),
    (29, "", None, None),
    (30, "COSTS", None, None, True),
    (31, "Venue (rent × 2 + deposit)", "=C11*2+C12", inr),
    (32, "Staff (× 2 months)", "=C13*2", inr),
    (33, "Materials (all kids)", "=C14*C6", inr),
    (34, "Snacks (all kids × 35 days)", "=C15*C6*35", inr),
    (35, "Per-child ops (all kids)", "=C16*C6", inr),
    (36, "Marketing", "=C17", inr),
    (37, "Insurance + ops", "=C18", inr),
    (38, "Subtotal", "=SUM(C31:C37)", inr),
    (39, "Buffer", "=C38*C19", inr),
    (40, "Total Costs", "=C38+C39", inr, True),
    (41, "", None, None),
    (42, "NET PROFIT / (LOSS)", "=C28-C40", inr, True),
    (43, "Margin %", "=IF(C28>0,C42/C28,0)", pct),
]

for item in results:
    row = item[0]
    label = item[1]
    formula = item[2] if len(item) > 2 else None
    fmt = item[3] if len(item) > 3 else None
    is_bold = item[4] if len(item) > 4 else False

    if is_bold and formula is None:
        ws.cell(row, 2, label).font = section_font
        for c in range(2, 5):
            ws.cell(row, c).border = Border(bottom=thick)
        continue

    ws.cell(row, 2, label).font = bold_black if is_bold else black_font
    if formula:
        cell = ws.cell(row, 3, formula)
        cell.font = bold_black if is_bold else black_font
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="center")

    if is_bold:
        for c in range(2, 5):
            ws.cell(row, c).border = double_border

# ── DASHBOARD CARDS (right side) ──
# Card 1: Revenue
ws.merge_cells("F5:G5")
ws.cell(5, 6, "TOTAL REVENUE").font = Font(color="6B7D8F", size=9, name="Arial", bold=True)
ws.cell(5, 6).alignment = Alignment(horizontal="center")
ws.merge_cells("F6:G6")
ws.cell(6, 6, "=C28").font = big_num
ws.cell(6, 6).number_format = inr
ws.cell(6, 6).alignment = Alignment(horizontal="center")

# Card 2: Costs
ws.merge_cells("F8:G8")
ws.cell(8, 6, "TOTAL COSTS").font = Font(color="6B7D8F", size=9, name="Arial", bold=True)
ws.cell(8, 6).alignment = Alignment(horizontal="center")
ws.merge_cells("F9:G9")
ws.cell(9, 6, "=C40").font = big_num
ws.cell(9, 6).number_format = inr
ws.cell(9, 6).alignment = Alignment(horizontal="center")

# Card 3: Profit
ws.merge_cells("F11:G11")
ws.cell(11, 6, "NET PROFIT").font = Font(color="6B7D8F", size=9, name="Arial", bold=True)
ws.cell(11, 6).alignment = Alignment(horizontal="center")
ws.merge_cells("F12:G12")
ws.cell(12, 6, "=C42").font = big_green
ws.cell(12, 6).number_format = inr
ws.cell(12, 6).alignment = Alignment(horizontal="center")

# Card 4: Margin
ws.merge_cells("F14:G14")
ws.cell(14, 6, "MARGIN").font = Font(color="6B7D8F", size=9, name="Arial", bold=True)
ws.cell(14, 6).alignment = Alignment(horizontal="center")
ws.merge_cells("F15:G15")
ws.cell(15, 6, "=C43").font = big_num
ws.cell(15, 6).number_format = pct
ws.cell(15, 6).alignment = Alignment(horizontal="center")

# Card 5: Break-even
ws.merge_cells("F17:G17")
ws.cell(17, 6, "BREAK-EVEN KIDS").font = Font(color="6B7D8F", size=9, name="Arial", bold=True)
ws.cell(17, 6).alignment = Alignment(horizontal="center")
ws.merge_cells("F18:G18")
# Break-even = Fixed costs / (Revenue per child - Variable cost per child)
# Fixed = venue + staff + marketing + insurance
# Variable per child = materials + snacks + ops
ws.cell(18, 6, '=ROUNDUP((C31+C32+C36+C37+((C31+C32+C36+C37)*C19)) / ((C7*C9+C8*(1-C9)) - (C14+C15*35+C16) - (C14+C15*35+C16)*C19), 0)').font = big_num
ws.cell(18, 6).number_format = num
ws.cell(18, 6).alignment = Alignment(horizontal="center")

# Card 6: Cost per child
ws.merge_cells("F20:G20")
ws.cell(20, 6, "COST PER CHILD").font = Font(color="6B7D8F", size=9, name="Arial", bold=True)
ws.cell(20, 6).alignment = Alignment(horizontal="center")
ws.merge_cells("F21:G21")
ws.cell(21, 6, "=IF(C6>0,C40/C6,0)").font = big_num
ws.cell(21, 6).number_format = inr
ws.cell(21, 6).alignment = Alignment(horizontal="center")

# ── SENSITIVITY TABLE ──
r = 45
ws.cell(r, 2, "SENSITIVITY: PROFIT BY NUMBER OF KIDS").font = section_font
for c in range(2, 8):
    ws.cell(r, c).border = Border(bottom=thick)

r = 46
headers = ["# Kids", "Revenue", "Total Costs", "Profit", "Margin"]
for c, h in enumerate(headers, 2):
    cell = ws.cell(r, c, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

kid_counts = [5, 8, 10, 12, 14, 16, 18, 20, 22, 24, 28, 30]
for i, kids in enumerate(kid_counts):
    r = 47 + i
    ws.cell(r, 2, kids).font = bold_black
    ws.cell(r, 2).alignment = Alignment(horizontal="center")

    # Revenue: kids * blended price
    ws.cell(r, 3, f"=B{r}*(C$7*C$9+C$8*(1-C$9))").font = black_font
    ws.cell(r, 3).number_format = inr
    ws.cell(r, 3).alignment = Alignment(horizontal="center")

    # Total costs: fixed + variable*kids + buffer
    fixed = "(C$11*2+C$12+C$13*2+C$17+C$18)"
    variable = f"(C$14+C$15*35+C$16)*B{r}"
    ws.cell(r, 4, f"=({fixed}+{variable})*(1+C$19)").font = black_font
    ws.cell(r, 4).number_format = inr
    ws.cell(r, 4).alignment = Alignment(horizontal="center")

    # Profit
    ws.cell(r, 5, f"=C{r}-D{r}").font = black_font
    ws.cell(r, 5).number_format = inr
    ws.cell(r, 5).alignment = Alignment(horizontal="center")

    # Margin
    ws.cell(r, 6, f"=IF(C{r}>0,E{r}/C{r},0)").font = black_font
    ws.cell(r, 6).number_format = pct
    ws.cell(r, 6).alignment = Alignment(horizontal="center")

# ── CHART ──
chart = BarChart()
chart.type = "col"
chart.title = "Revenue vs Costs by # of Kids"
chart.y_axis.title = "Amount (₹)"
chart.x_axis.title = "Number of Kids"
chart.style = 10
chart.width = 28
chart.height = 14

cats = Reference(ws, min_col=2, min_row=47, max_row=58)
rev_data = Reference(ws, min_col=3, min_row=46, max_row=58)
cost_data = Reference(ws, min_col=4, min_row=46, max_row=58)

chart.add_data(rev_data, titles_from_data=True)
chart.add_data(cost_data, titles_from_data=True)
chart.set_categories(cats)

chart.series[0].graphicalProperties.solidFill = "2B5797"
chart.series[1].graphicalProperties.solidFill = "E05A3A"

# Add profit as line
line = LineChart()
profit_data = Reference(ws, min_col=5, min_row=46, max_row=58)
line.add_data(profit_data, titles_from_data=True)
line.series[0].graphicalProperties.line.solidFill = "3A8C6E"
line.series[0].graphicalProperties.line.width = 25000
line.y_axis.axId = 200

chart.y_axis.crosses = "min"
chart += line

ws.add_chart(chart, "B60")

output = "/Users/garimatyagi/Coding/summer-camp/Summer_Club_Budget.xlsx"
wb.save(output)
print(f"Saved with Simulator sheet to {output}")
