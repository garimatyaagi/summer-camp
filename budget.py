from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
blue_font = Font(color="0000FF", size=11, name="Arial")
black_font = Font(color="000000", size=11, name="Arial")
bold_black = Font(color="000000", size=11, name="Arial", bold=True)
bold_blue = Font(color="0000FF", size=11, name="Arial", bold=True)
header_font = Font(color="FFFFFF", size=12, name="Arial", bold=True)
section_font = Font(color="2B5797", size=13, name="Arial", bold=True)
title_font = Font(color="2B5797", size=16, name="Arial", bold=True)

header_fill = PatternFill("solid", fgColor="2B5797")
light_fill = PatternFill("solid", fgColor="E8EDF3")
yellow_fill = PatternFill("solid", fgColor="FFFF00")
green_fill = PatternFill("solid", fgColor="E2EFDA")
red_fill = PatternFill("solid", fgColor="FCE4EC")
white_fill = PatternFill("solid", fgColor="FFFFFF")

inr = '₹#,##0;(₹#,##0);"-"'
inr_bold = '₹#,##0;(₹#,##0);"-"'
pct = '0.0%'
num = '#,##0'
thin = Side(style="thin", color="D4DCE6")
border = Border(bottom=thin)
thick_border = Border(bottom=Side(style="medium", color="2B5797"))

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

def style_section(ws, row, cols, text):
    ws.cell(row=row, column=1, value=text).font = section_font
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).border = Border(bottom=Side(style="medium", color="2B5797"))

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════
# SHEET 1: REVENUE SCENARIOS
# ═══════════════════════════════════════
ws1 = wb.active
ws1.title = "Revenue"
set_widths(ws1, [35, 18, 18, 18, 18])

ws1.cell(1, 1, "THE SUMMER CLUB — REVENUE SCENARIOS").font = title_font
ws1.cell(2, 1, "7-week programme · April 14 – May 30, 2026").font = Font(color="6B7D8F", size=11, name="Arial")

# Assumptions
r = 4
style_section(ws1, r, 5, "ASSUMPTIONS")
r = 5
for label, val, fmt, col_font in [
    ("Early bird price (per child)", 21000, inr, blue_font),
    ("Regular price (per child)", 24500, inr, blue_font),
    ("Early bird mix", 0.6, pct, blue_font),
    ("Regular mix", None, pct, black_font),
]:
    r += 1
    ws1.cell(r, 1, label).font = black_font
    if val is not None:
        ws1.cell(r, 2, val).font = col_font
        ws1.cell(r, 2).number_format = fmt
        if fmt == pct:
            ws1.cell(r, 2).fill = yellow_fill
    else:
        ws1.cell(r, 2, "=1-B8").font = black_font
        ws1.cell(r, 2).number_format = pct

# Scenarios
r = 11
style_section(ws1, r, 5, "SCENARIOS")
r = 12
headers = ["", "Scenario A\n(Conservative)", "Scenario B\n(Target)", "Scenario C\n(Stretch)"]
for c, h in enumerate(headers, 1):
    cell = ws1.cell(r, c, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[r].height = 35

data = [
    ("Total kids", [12, 18, 24], num, blue_font),
    ("Early bird kids", ["=B13*B8", "=C13*B8", "=D13*B8"], num, black_font),
    ("Regular price kids", ["=B13-B14", "=C13-C14", "=D13-D14"], num, black_font),
    ("", [], None, None),
    ("Early bird revenue", ["=B14*B6", "=C14*B6", "=D14*B6"], inr, black_font),
    ("Regular revenue", ["=B15*B7", "=C15*B7", "=D15*B7"], inr, black_font),
    ("", [], None, None),
    ("TOTAL REVENUE", ["=B17+B18", "=C17+C18", "=D17+D18"], inr, bold_black),
    ("Revenue per child", ["=B20/B13", "=C20/C13", "=D20/D13"], inr, black_font),
]

for i, (label, vals, fmt, font) in enumerate(data):
    r = 13 + i
    ws1.cell(r, 1, label).font = font if font else black_font
    if label == "TOTAL REVENUE":
        ws1.cell(r, 1).font = bold_black
        for c in range(1, 5):
            ws1.cell(r, c).border = Border(top=Side(style="medium", color="2B5797"), bottom=Side(style="double", color="2B5797"))
    for c, v in enumerate(vals, 2):
        cell = ws1.cell(r, c, v)
        cell.font = font
        if fmt:
            cell.number_format = fmt
        cell.alignment = Alignment(horizontal="center")

# ═══════════════════════════════════════
# SHEET 2: DETAILED COSTS
# ═══════════════════════════════════════
ws2 = wb.create_sheet("Costs")
set_widths(ws2, [38, 18, 18, 15, 30])

ws2.cell(1, 1, "THE SUMMER CLUB — DETAILED COSTS").font = title_font
ws2.cell(2, 1, "Based on target batch of 18 kids").font = Font(color="6B7D8F", size=11, name="Arial")

headers = ["Item", "Monthly / Unit", "Duration / Qty", "Total", "Notes"]
r = 4
for c, h in enumerate(headers, 1):
    cell = ws2.cell(r, c, h)
    cell.font = header_font
    cell.fill = header_fill

costs = [
    ("VENUE", None, None, None, None, True),
    ("Activity space rental", 20000, "2 months", "=B6*2", "Community hall / co-working space, mornings only"),
    ("Security deposit (refundable)", 20000, "1", "=B7", "Typically 1 month rent"),
    ("", None, None, None, None, False),
    ("STAFF", None, None, None, None, True),
    ("Lead instructor #1", 20000, "2 months", "=B10*2", "Art/science background, full-time for camp"),
    ("Lead instructor #2", 18000, "2 months", "=B11*2", "Sports/movement + cooking"),
    ("Helper / assistant", 12000, "2 months", "=B12*2", "Setup, cleanup, snack prep, supervision"),
    ("First aid trained person", 5000, "2 months", "=B13*2", "Part-time, on-call during camp hours"),
    ("", None, None, None, None, False),
    ("MATERIALS (per child per week)", None, None, None, None, True),
    ("Week 1: Art supplies", 400, 18, "=B16*C16", "Clay, paint, canvas, brushes, tie-dye kit"),
    ("Week 2: Science kits", 350, 18, "=B17*C17", "Baking soda, magnets, slime ingredients, seed kits"),
    ("Week 3: Cooking ingredients", 300, 18, "=B18*C18", "Fruits, oats, honey, lemonade supplies"),
    ("Week 4: Nature supplies", 200, 18, "=B19*C19", "Bug hotel materials, pressed flower supplies, journals"),
    ("Week 5: Robot building", 450, 18, "=B20*C20", "Cardboard, foil, LEDs, tubes, coding cards"),
    ("Week 6: Market supplies", 350, 18, "=B21*C21", "Craft materials for products, signage, play money"),
    ("Week 7: Exhibition", 500, 18, "=B22*C22", "Frames, labels, certificates, scrapbooks, t-shirts"),
    ("", None, None, None, None, False),
    ("FOOD & SNACKS", None, None, None, None, True),
    ("Daily snacks (fruit + juice)", 30, "18 kids × 35 days", "=B25*18*35", "₹30/child/day"),
    ("", None, None, None, None, False),
    ("MARKETING", None, None, None, None, True),
    ("Flyers & posters printing", 3000, 1, "=B28", "200 flyers + 20 posters"),
    ("Instagram promoted posts", 5000, 1, "=B29", "Optional — boost 2-3 posts"),
    ("", None, None, None, None, False),
    ("OPERATIONS", None, None, None, None, True),
    ("Insurance", 8000, 1, "=B32", "Group activity insurance for 7 weeks"),
    ("First aid kit + supplies", 3000, 1, "=B33", "Comprehensive kit"),
    ("Certificates printing", 150, 18, "=B34*C34", "Custom designed, heavy stock"),
    ("Name tags + lanyards", 100, 18, "=B35*C35", "Printed, durable"),
    ("Camp t-shirts", 350, 18, "=B36*C36", "Custom printed, given in Week 7"),
    ("Scrapbooks", 200, 18, "=B37*C37", "Blank hardcover, given in Week 7"),
    ("", None, None, None, None, False),
    ("MISCELLANEOUS", None, None, None, None, True),
    ("Buffer (10% of above)", None, None, "=SUM(D6:D37)*0.1", "Unexpected costs, replacements, extras"),
]

r = 4
for item in costs:
    r += 1
    label, monthly, dur, total, notes, is_section = item[0], item[1], item[2], item[3], item[4], item[5] if len(item) > 5 else False
    if is_section:
        style_section(ws2, r, 5, label)
        continue
    if not label:
        continue
    ws2.cell(r, 1, label).font = black_font
    if monthly is not None:
        ws2.cell(r, 2, monthly).font = blue_font
        ws2.cell(r, 2).number_format = inr
        ws2.cell(r, 2).fill = yellow_fill
    if dur is not None:
        ws2.cell(r, 3, dur).font = blue_font
        ws2.cell(r, 3).alignment = Alignment(horizontal="center")
    if total is not None:
        ws2.cell(r, 4, total).font = black_font
        ws2.cell(r, 4).number_format = inr
    if notes:
        ws2.cell(r, 5, notes).font = Font(color="6B7D8F", size=10, name="Arial", italic=True)

# Total costs
r += 2
ws2.cell(r, 1, "TOTAL COSTS").font = bold_black
ws2.cell(r, 4, "=SUM(D6:D37)+D40").font = bold_black
ws2.cell(r, 4).number_format = inr
for c in range(1, 5):
    ws2.cell(r, c).border = Border(top=Side(style="medium", color="2B5797"), bottom=Side(style="double", color="2B5797"))

# ═══════════════════════════════════════
# SHEET 3: P&L SUMMARY
# ═══════════════════════════════════════
ws3 = wb.create_sheet("P&L Summary")
set_widths(ws3, [35, 18, 18, 18])

ws3.cell(1, 1, "THE SUMMER CLUB — P&L SUMMARY").font = title_font

r = 3
headers = ["", "Conservative\n(12 kids)", "Target\n(18 kids)", "Stretch\n(24 kids)"]
for c, h in enumerate(headers, 1):
    cell = ws3.cell(r, c, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws3.row_dimensions[r].height = 35

pnl_data = [
    ("Revenue", ["=Revenue!B20", "=Revenue!C20", "=Revenue!D20"], inr, bold_black, None),
    ("", [], None, None, None),
    ("Fixed Costs", None, None, None, None),
    ("  Venue", ["=Costs!D6+Costs!D7", "=Costs!D6+Costs!D7", "=Costs!D6+Costs!D7"], inr, black_font, None),
    ("  Staff", ["=Costs!D10+Costs!D11+Costs!D12+Costs!D13", "=Costs!D10+Costs!D11+Costs!D12+Costs!D13", "=Costs!D10+Costs!D11+Costs!D12+Costs!D13"], inr, black_font, None),
    ("  Marketing", ["=Costs!D28+Costs!D29", "=Costs!D28+Costs!D29", "=Costs!D28+Costs!D29"], inr, black_font, None),
    ("  Insurance + ops (fixed)", ["=Costs!D32+Costs!D33", "=Costs!D32+Costs!D33", "=Costs!D32+Costs!D33"], inr, black_font, None),
    ("Total Fixed Costs", ["=SUM(B6:B9)", "=SUM(C6:C9)", "=SUM(D6:D9)"], inr, bold_black, None),
    ("", [], None, None, None),
    ("Variable Costs (per child)", None, None, None, None),
    ("  Materials (7 weeks)", ["=SUM(Costs!D16:Costs!D22)/18*12", "=SUM(Costs!D16:Costs!D22)", "=SUM(Costs!D16:Costs!D22)/18*24"], inr, black_font, None),
    ("  Snacks (35 days)", ["=30*12*35", "=30*18*35", "=30*24*35"], inr, black_font, None),
    ("  T-shirts + scrapbooks + certs + tags", ["=(350+200+150+100)*12", "=(350+200+150+100)*18", "=(350+200+150+100)*24"], inr, black_font, None),
    ("Total Variable Costs", ["=SUM(B13:B15)", "=SUM(C13:C15)", "=SUM(D13:D15)"], inr, bold_black, None),
    ("", [], None, None, None),
    ("Buffer (10%)", ["=(B10+B16)*0.1", "=(C10+C16)*0.1", "=(D10+D16)*0.1"], inr, black_font, None),
    ("", [], None, None, None),
    ("TOTAL COSTS", ["=B10+B16+B18", "=C10+C16+C18", "=D10+D16+D18"], inr, bold_black, thick_border),
    ("", [], None, None, None),
    ("NET PROFIT / (LOSS)", ["=B4-B20", "=C4-C20", "=D4-D20"], inr, bold_black, None),
    ("Margin %", ["=IF(B4>0,B22/B4,0)", "=IF(C4>0,C22/C4,0)", "=IF(D4>0,D22/D4,0)"], pct, black_font, None),
    ("", [], None, None, None),
    ("PER-CHILD ECONOMICS", None, None, None, None),
    ("Revenue per child", ["=Revenue!B21", "=Revenue!C21", "=Revenue!D21"], inr, black_font, None),
    ("Cost per child", ["=B20/12", "=C20/18", "=D20/24"], inr, black_font, None),
    ("Margin per child", ["=B26-B27", "=C26-C27", "=D26-D27"], inr, black_font, None),
    ("", [], None, None, None),
    ("BREAK-EVEN", None, None, None, None),
    ("Break-even # of kids", ["=ROUNDUP(B10/(Revenue!B21-B16/12),0)", "=ROUNDUP(C10/(Revenue!C21-C16/18),0)", "=ROUNDUP(D10/(Revenue!D21-D16/24),0)"], num, bold_blue, yellow_fill),
]

for i, (label, vals, fmt, font, special_border) in enumerate(pnl_data):
    r = 4 + i
    ws3.cell(r, 1, label).font = font if font else black_font
    if label in ("Fixed Costs", "Variable Costs (per child)", "PER-CHILD ECONOMICS", "BREAK-EVEN"):
        style_section(ws3, r, 4, label)
        continue
    if label in ("TOTAL COSTS", "NET PROFIT / (LOSS)"):
        for c in range(1, 5):
            ws3.cell(r, c).border = Border(top=Side(style="medium", color="2B5797"), bottom=Side(style="double", color="2B5797"))
    if label == "NET PROFIT / (LOSS)":
        for c in range(2, 5):
            ws3.cell(r, c).fill = green_fill
    if vals:
        for c, v in enumerate(vals, 2):
            cell = ws3.cell(r, c, v)
            cell.font = font
            if fmt:
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal="center")
            if special_border:
                cell.fill = special_border if isinstance(special_border, PatternFill) else white_fill

# ═══════════════════════════════════════
# SHEET 4: CASH FLOW
# ═══════════════════════════════════════
ws4 = wb.create_sheet("Cash Flow")
set_widths(ws4, [35, 18, 18, 18, 18, 18])

ws4.cell(1, 1, "THE SUMMER CLUB — CASH FLOW").font = title_font
ws4.cell(2, 1, "Based on target scenario (18 kids)").font = Font(color="6B7D8F", size=11, name="Arial")

r = 4
headers = ["", "March\n(Pre-camp)", "April\n(Weeks 1-3)", "May\n(Weeks 4-7)", "June\n(Wrap-up)", "TOTAL"]
for c, h in enumerate(headers, 1):
    cell = ws4.cell(r, c, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws4.row_dimensions[r].height = 35

cf_data = [
    ("MONEY IN", None, True),
    ("Registrations received", [0, 15, 3, 0], True),
    ("Registration revenue", ["=B6*Revenue!C21", "=C6*Revenue!C21", "=D6*Revenue!C21", "=E6*Revenue!C21"], False),
    ("Total Inflow", ["=B7", "=C7", "=D7", "=E7"], False, True),
    ("", [], False),
    ("MONEY OUT", None, True),
    ("Venue deposit", [20000, 0, 0, -20000], False),
    ("Venue rent", [0, 20000, 20000, 0], False),
    ("Staff salaries", [0, 55000, 55000, 0], False),
    ("Materials — Weeks 1-3", [15000, 34650, 0, 0], False),
    ("Materials — Weeks 4-7", [0, 0, 46350, 0], False),
    ("Snacks", [0, 8100, 10800, 0], False),
    ("Marketing", [5000, 3000, 0, 0], False),
    ("Insurance + ops", [11000, 3500, 0, 0], False),
    ("T-shirts, certs, scrapbooks", [0, 0, 14400, 0], False),
    ("Buffer", [2000, 5000, 5000, 0], False),
    ("Total Outflow", ["=SUM(B11:B20)", "=SUM(C11:C20)", "=SUM(D11:D20)", "=SUM(E11:E20)"], False, True),
    ("", [], False),
    ("NET CASH FLOW", ["=B8-B21", "=C8-C21", "=D8-D21", "=E8-E21"], False, True),
    ("CUMULATIVE BALANCE", ["=B23", "=B24+C23", "=C24+D23", "=D24+E23"], False, True),
]

for i, item in enumerate(cf_data):
    label = item[0]
    vals = item[1] if len(item) > 1 else []
    is_section = item[2] if len(item) > 2 else False
    is_total = item[3] if len(item) > 3 else False
    r = 5 + i
    
    if is_section and vals is None:
        style_section(ws4, r, 6, label)
        continue
    
    ws4.cell(r, 1, label).font = bold_black if is_total else black_font
    
    if vals:
        for c, v in enumerate(vals, 2):
            cell = ws4.cell(r, c, v)
            cell.font = bold_black if is_total else (blue_font if isinstance(v, (int, float)) else black_font)
            cell.number_format = inr
            cell.alignment = Alignment(horizontal="center")
            if isinstance(v, (int, float)):
                cell.fill = yellow_fill
        # Total column
        if len(vals) == 4:
            cell = ws4.cell(r, 6, f"=SUM(B{r}:E{r})")
            cell.font = bold_black if is_total else black_font
            cell.number_format = inr
            cell.alignment = Alignment(horizontal="center")
    
    if is_total:
        for c in range(1, 7):
            ws4.cell(r, c).border = Border(top=Side(style="medium", color="2B5797"), bottom=Side(style="double", color="2B5797"))
    
    if label == "CUMULATIVE BALANCE":
        for c in range(2, 6):
            ws4.cell(r, c).fill = green_fill

output = "/Users/garimatyagi/Coding/summer-camp/Summer_Club_Budget.xlsx"
wb.save(output)
print(f"Saved to {output}")
